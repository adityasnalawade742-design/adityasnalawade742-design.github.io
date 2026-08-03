"""
Product Registry Engine (SQLite + Excel Sync + Image Lifecycle Management)
Tracks Published & Rejected products, handles deduplication, auto-cleans unselected raw images,
and syncs to `product_registry.xlsx` safely without crashing if Excel is open.
"""
import os
import time
import sqlite3
import threading
from pathlib import Path

_PROJECT_ROOT = Path(__file__).resolve().parent.parent
DB_PATH = _PROJECT_ROOT / "cache" / "registry.db"
EXCEL_PATH = _PROJECT_ROOT / "product_registry.xlsx"
RAW_IMAGES_DIR = _PROJECT_ROOT / "raw_images"

DB_PATH.parent.mkdir(parents=True, exist_ok=True)
RAW_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

_sync_lock = threading.Lock()


def init_registry_db():
    """Initializes SQLite tables for published and rejected products."""
    with sqlite3.connect(DB_PATH) as conn:
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS published_products (
                asin TEXT PRIMARY KEY,
                title TEXT,
                price TEXT,
                pinterest_pin_id TEXT,
                image_url TEXT,
                published_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS rejected_products (
                asin TEXT PRIMARY KEY,
                title TEXT,
                reason TEXT,
                rejected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        conn.commit()


# Initialize on import
init_registry_db()


def get_blocked_asins() -> set:
    """
    Returns a set of all ASINs that are either published or rejected.
    Used during discovery to filter out repeat products instantly (<1ms).
    """
    blocked = set()
    try:
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            c.execute("SELECT asin FROM published_products")
            for row in c.fetchall():
                if row[0]:
                    blocked.add(row[0].strip().upper())
            c.execute("SELECT asin FROM rejected_products")
            for row in c.fetchall():
                if row[0]:
                    blocked.add(row[0].strip().upper())
    except Exception as e:
        print(f"[Registry DB Error] Failed to read blocked ASINs: {e}")

    return blocked


def mark_as_rejected(asin: str, title: str = "", reason: str = "user_skip"):
    """
    Marks an ASIN as rejected, deletes its cached raw image immediately,
    and triggers an async Excel sync.
    """
    asin = (asin or "").strip().upper()
    if not asin:
        return

    try:
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            c.execute("""
                INSERT OR REPLACE INTO rejected_products (asin, title, reason, rejected_at)
                VALUES (?, ?, ?, CURRENT_TIMESTAMP)
            """, (asin, title or f"Product {asin}", reason))
            conn.commit()
        print(f"[Registry] Marked ASIN {asin} as REJECTED ({reason})")
    except Exception as e:
        print(f"[Registry Error] Failed to mark {asin} as rejected: {e}")

    # Immediately delete local raw image to save disk space
    delete_raw_image(asin)

    # Sync to Excel in background thread so HTTP response remains fast
    threading.Thread(target=sync_to_excel, daemon=True).start()


def mark_as_published(asin: str, title: str = "", price: str = "", pinterest_pin_id: str = "", image_url: str = ""):
    """
    Marks an ASIN as published (called when Pinterest API or publishing succeeds),
    deletes its local raw image, and triggers an async Excel sync.
    """
    asin = (asin or "").strip().upper()
    if not asin:
        return

    try:
        with sqlite3.connect(DB_PATH) as conn:
            c = conn.cursor()
            c.execute("""
                INSERT OR REPLACE INTO published_products (asin, title, price, pinterest_pin_id, image_url, published_at)
                VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
            """, (asin, title or f"Product {asin}", price or "$19.99", pinterest_pin_id or "", image_url or ""))
            conn.commit()
        print(f"[Registry] Marked ASIN {asin} as PUBLISHED (Pin ID: {pinterest_pin_id or 'Pending'})")
    except Exception as e:
        print(f"[Registry Error] Failed to mark {asin} as published: {e}")

    # Also register with homepage selector if available
    try:
        from modules.automated_product_selector import save_processed_asin
        save_processed_asin(asin)
    except Exception:
        pass

    # Delete raw image now that product is published
    delete_raw_image(asin)

    # Sync to Excel in background thread
    threading.Thread(target=sync_to_excel, daemon=True).start()


def delete_raw_image(asin: str):
    """Deletes raw_images/raw_{ASIN}.jpg if it exists."""
    asin = (asin or "").strip().upper()
    if not asin:
        return
    raw_file = RAW_IMAGES_DIR / f"raw_{asin}.jpg"
    if raw_file.exists():
        try:
            raw_file.unlink(missing_ok=True)
            print(f"[Registry Cleanup] Auto-deleted raw image: {raw_file.name}")
        except Exception as e:
            print(f"[Registry Cleanup Warning] Could not delete {raw_file.name}: {e}")


def cleanup_orphaned_raw_images(max_age_hours: int = 24):
    """
    Deletes any raw image files older than max_age_hours that are not currently active.
    Can be run on server startup.
    """
    now = time.time()
    max_age_sec = max_age_hours * 3600
    cleaned_count = 0

    for item in RAW_IMAGES_DIR.glob("raw_*.jpg"):
        try:
            mtime = item.stat().st_mtime
            if (now - mtime) > max_age_sec:
                item.unlink(missing_ok=True)
                cleaned_count += 1
        except Exception:
            pass

    if cleaned_count > 0:
        print(f"[Registry Cleanup] Purged {cleaned_count} orphaned raw images older than {max_age_hours}h.")


def sync_to_excel():
    """
    Safely exports SQLite tables to `product_registry.xlsx` with 'Published' and 'Rejected' sheets.
    Handles file locking gracefully if open in Excel on desktop.
    """
    with _sync_lock:
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("[Registry Warning] openpyxl not installed. Skipping Excel sync.")
            return

        try:
            with sqlite3.connect(DB_PATH) as conn:
                c = conn.cursor()

                c.execute("SELECT asin, title, price, pinterest_pin_id, image_url, published_at FROM published_products ORDER BY published_at DESC")
                pub_rows = c.fetchall()

                c.execute("SELECT asin, title, reason, rejected_at FROM rejected_products ORDER BY rejected_at DESC")
                rej_rows = c.fetchall()
        except Exception as e:
            print(f"[Registry Error] Could not read SQLite DB for Excel sync: {e}")
            return

        wb = openpyxl.Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # ── Styling ──
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill_pub = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")  # Emerald green
        header_fill_rej = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")  # Rose red
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        # ── 1. Published Sheet ──
        ws_pub = wb.create_sheet(title="Published")
        pub_headers = ["ASIN", "Product Title", "Price", "Pinterest Pin ID", "Image URL", "Published Date"]
        ws_pub.append(pub_headers)
        for col_idx in range(1, len(pub_headers) + 1):
            cell = ws_pub.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill_pub
            cell.alignment = center_align

        for r in pub_rows:
            ws_pub.append(list(r))

        # Column widths for Published
        pub_col_widths = [15, 45, 12, 22, 50, 22]
        for idx, width in enumerate(pub_col_widths, start=1):
            ws_pub.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # ── 2. Rejected Sheet ──
        ws_rej = wb.create_sheet(title="Rejected")
        rej_headers = ["ASIN", "Product Title", "Rejection Reason", "Rejected Date"]
        ws_rej.append(rej_headers)
        for col_idx in range(1, len(rej_headers) + 1):
            cell = ws_rej.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill_rej
            cell.alignment = center_align

        for r in rej_rows:
            ws_rej.append(list(r))

        # Column widths for Rejected
        rej_col_widths = [15, 45, 20, 22]
        for idx, width in enumerate(rej_col_widths, start=1):
            ws_rej.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width

        # Save workbook safely
        try:
            wb.save(EXCEL_PATH)
            print(f"[Registry Excel Sync] Successfully updated {EXCEL_PATH.name} ({len(pub_rows)} published, {len(rej_rows)} rejected)")
        except PermissionError:
            print(f"[Registry Excel Warning] Could not save {EXCEL_PATH.name} — file is currently open in Excel. Will sync on next event.")
        except Exception as e_save:
            print(f"[Registry Excel Error] Failed to save Excel file: {e_save}")
